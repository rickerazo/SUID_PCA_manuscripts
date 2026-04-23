'''
Dr Ricardo Erazo
Fall 2023
Seattle Children's Hospital
Seattle Children's Research Institute 
Center for Integrative Brain Research
Jan Ramirez, PhD

The University of Auckland,
Department of Paediatrics, Youth and Health
Ed Mitchel, PhD



THIS SCRIPT PREPROCESS THE DATABASE.
HERE WE IMPLEMENT SEVERAL FILTERS
WE ACCOUNT FOR BLANK AND UNKNOWN DATA.

The output is a csv file that includes the original-modified data
appended additional variables fruit of pre-analysis

'''


# import sys
import os
# ninja_dir = os.getcwd()
# ninja_directory=ninja_dir[0:-4]+'CDC_data_project/'
# sys.path.append(ninja_directory)
# from ninja_functions import count_individuals
# from SUID_functions import
####
from openpyxl import load_workbook
import pandas as pd
from pyreadstat import read_sav
import numpy as np


import matplotlib.pyplot as plt
import seaborn as sns

os.makedirs('output_figures/',exist_ok=True)
os.makedirs('output_data/',exist_ok=True)


## include ninja functions
# data_dir = '/active/ramirez_j/ramirezlab/SIDS MSFT collaboration/NCFRP Data/'
data_dir = 'data/'
df,meta = read_sav(data_dir+'SeattleChildrensNCFRP.sav')
# df,meta = read_sav(data_dir+'SeattleChildrensNCFRP.sav')
							 # SeattleChildrensNCFRP.sav
### documentation:
workbook = load_workbook(data_dir+'Codebook_Version5-1_031122.xlsx')
code_sheets = workbook.sheetnames
## Ed's Code:
eds_handbook = pd.read_excel('from_Ed/Eds_ICD_codes.xlsx')
plt.rcParams['font.size']=17
# ######## COMMENT IF WANT TO UNSEE THIS INFORMATION:
# print('List of sheets in code book:')
# for x in code_sheets: print(x)
# print('\nList of columns in dataset: ')
# for x in df.columns.tolist(): print(x)

########### functions used multiple times:
import selection

def select_sample(tab):
	date_min=2000
	date_max=2019

	_max = tab.index[tab['INFdodyear']<=2019]
	dt = tab.iloc[_max]
	_min = dt.index[dt['INFdodyear']>=2008]
	tab = tab.iloc[_min].copy()
	return tab

###### GESTATIONAL AGE AND BIRTH WEIGHT
g_age = np.array([24 , 25 , 26 , 27 , 28 , 29 , 30 , 31 , 32 , 33 , 34 , 35 , 36 , 37 , 38 , 39 , 40 , 41,	42, 43])
ct_99 = np.array([820, 957,1110,1278,1461,1658,1869,2091,2324,2564,2809,3056,3301,3540,3770,3987,2186,4365,4365,4365])
ct_97 = np.array([786, 918,1064,1225,1401,1590,1792,2005,2228,2459,2694,2930,3165,3395,3615,3823,4014,4185,4185,4185])
ct_95 = np.array([768, 897,1040,1198,1369,1554,1751,1960,2178,2403,2632,2864,3093,3318,3533,3736,3923,4090,4090,4090])
ct_90 = np.array([741, 865,1003,1155,1320,1498,1689,1890,2100,2317,2538,2761,2983,3199,3407,3603,3783,3944,3944,3944])
ct_75 = np.array([695, 812, 941,1083,1238,1405,1584,1773,1970,2173,2381,2590,2798,3001,3196,3380,3549,3700,3700,3700])
ct_50 = np.array([644, 752, 872,1003,1147,1302,1468,1643,1825,2014,2206,2400,2593,2781,2961,3132,3288,3428,3428,3428])
ct_25 = np.array([593, 692, 803, 924,1057,1119,1352,1513,1681,1854,2032,2210,2387,2561,2727,2884,3028,3157,3157,3157])
ct_10 = np.array([547, 639, 741, 853, 975,1106,1247,1395,1551,1711,1874,2039,2203,2362,2516,2660,2794,2913,2913,2913])
ct_05 = np.array([520, 607, 703, 810, 926,1051,1184,1325,1473,1625,1780,1937,2092,2244,2390,2527,2653,2766,2766,2766])
ct_03 = np.array([502, 586, 679, 782, 894,1015,1144,1280,1422,1569,1719,1870,2020,2167,2308,2440,2562,2671,2671,2671])
ct_01 = np.array([468, 547, 634, 730, 834, 947,1067,1194,1327,1464,1604,1745,1885,2021,2153,2276,2390,2492,2492,2492])

tab = pd.DataFrame({'99%':ct_99,'97%':ct_97,'95%':ct_95,'90%':ct_90,'75%':ct_75,'mean':ct_50,'25%':ct_25,'10%':ct_10,'5%':ct_05,'1%':ct_01}, index=g_age)

##########	INF 

df.loc[df.index[df['INFethnic']==0],'INFethnic']=9
df.loc[df.index[df['INFsex']==0], 'INFsex'] = 9

df.loc[df.index[df['INF3multbirth']!=1], 'INF3multbirth']=2
############################# ## Infant birthweight

df.loc[df[df['INFagecat']==2].index, 'age_decoder'] = 30.	#### estimate infant age in days, if the  data is in months: multiply by 30.
df.loc[df[df['INFagecat']==3].index, 'age_decoder'] = 1.	#### if data is already encoded in days, pass number 1 for multiplication.

df['INFageDays'] = df['INFage'].astype(int)*df['age_decoder']	## actually compute the infants' ages


df.loc[np.nonzero(df['INFageDays']<30)[0], 'INFageMonths'] = 0
df.loc[df['INFageDays']<360, 'INFageMonths'] = 11
df.loc[df['INFageDays']<330, 'INFageMonths'] = 10
df.loc[df['INFageDays']<300, 'INFageMonths'] = 9
df.loc[df['INFageDays']<270, 'INFageMonths'] = 8
df.loc[df['INFageDays']<240, 'INFageMonths'] = 7
df.loc[df['INFageDays']<210, 'INFageMonths'] = 6
df.loc[df['INFageDays']<180, 'INFageMonths'] = 5
df.loc[df['INFageDays']<150, 'INFageMonths'] = 4
df.loc[df['INFageDays']<120, 'INFageMonths'] = 3
df.loc[df['INFageDays']<90, 'INFageMonths'] = 2
df.loc[df['INFageDays']<60, 'INFageMonths'] = 1
df.loc[df['INFageDays']<30, 'INFageMonths'] = 0

#################################################################################### INFANT AGE AND WEIGHT PREPROCESSING
'''
Systematically remove outliers using reference table. Infant gestation age is the independent variable - criterion to remove outliers.
1. loop through reference table: Use it to identify individuals with the same gestational age.
2. pooling all individuals of the same gestational age, compute mean and sd
3. compute 99 percentile + sd
4. compute 1 percentile + sd
5. find individuals that are beyond the ceiling threshold

'''
df['INF3biwt']=df['INF3birthwtgrams'].copy()
for i in range(tab.index[0],tab.index[-1]+1):
	wtclass = df.iloc[df.index[df['INF3gestage']==i]]
	# class_mn = np.mean(wtclass['INF3birthwtgrams'])

	# class_sd = np.std(wtclass['INF3birthwtgrams'])	
	# threshold_ceil = tab.loc[i, '99%']+class_sd
	# threshold_flor = tab.loc[i, '1%']-class_sd
	# print(f'\n{i},  {threshold_flor:.2f},{threshold_ceil:.2f}')

	# beyond = wtclass.loc[wtclass.index[wtclass['INF3birthwtgrams']>threshold_ceil], 'INF3birthwtgrams']
	# below = wtclass.loc[wtclass.index[wtclass['INF3birthwtgrams']<threshold_flor], 'INF3birthwtgrams']
	# print(len(below),len(beyond))

	class_sd = (tab.loc[i, '90%'] - tab.loc[i,'10%'])/2.698
	threshold_ceil = tab.loc[i, '90%']+class_sd
	threshold_flor = tab.loc[i, '10%']-class_sd
	# print(f'{i}, {threshold_flor:.2f},{threshold_ceil:.2f}')

	beyond = wtclass.loc[wtclass.index[wtclass['INF3birthwtgrams']>threshold_ceil], 'INF3birthwtgrams']
	below = wtclass.loc[wtclass.index[wtclass['INF3birthwtgrams']<threshold_flor], 'INF3birthwtgrams']
	# print(len(below),len(beyond))

	df.loc[beyond.index, 'INF3biwt'] = np.nan
	df.loc[below.index, 'INF3biwt'] = np.nan

df.loc[df.index[df['INF3gestage']>tab.index[-1]], 'INF3biwt'] = np.nan

###################################################################
### Problematic variables

df.loc[df.index[df['INF3birthwtgrams']==0], 'INF3birthwtunk'] = 1
df.loc[df.index[df['INF3birthwtgrams']==0], 'INF3birthwtgrams'] = np.nan
df.loc[pd.isna(df['INF3birthwtgrams']),'INF3birthwtunk'] = 1

df['INF3birthwtBIN'] = 0
df.loc[pd.isna(df['INF3birthwtgrams']), 'INF3birthwtBIN'] = 1

########################################################

# ######### clean up multbirth
# ############## should this me 0 or 1?

### 	THERE ARE MISSING DATA:
### 	MULTBIRTH AND MULTNUM AREN'T CONSISTENTLY REPORTED
### 	APPLYING HIERARCHIES:
### 		1. PRIORITY TO POSITIVE RESPONSES TO MULTI BIRTH
### 		2. we'll assume that every positive multibirth had at least 2 infants delivered.
### 		3. Fill-in missing data: blanks.
### 		4. Correct infmultnum=1 when multibirth=1, it should be at least 2.

##### If there was a positive answer to MULTBIRTH, then MULTNUM is AT LEAST 2:
p1 = df.iloc[df.index[df['INF3multbirth']==1]]
idx= p1.index[p1['INF3multnum'].isna()]
df.loc[idx, 'INF3multnum']=2
df.loc[df.index[df['INF3multnum']==1], 'INF3multnum']=2
# conso = df.iloc[df.index[df['INF3multnum'].isna()]]

### 83 subjects responded yes to multibirth, but responded to 1 in multnum. 
## We are going to fix that, by modifying the multnum to 2

# df.loc[df.index[df['INF3multnum']==1], 'INF3multnum'] = 0
# df.index[df['INF3multbirth']==2]
# df.loc[df.index[df['INF3multbirth']==1], 'INF3multnum']=2


####################################################################################
####################################################################################
####################################################################################
####################################################################################
df.loc[df.index[df['INF3prenatalcare']==0], 'INF3prenatalcare'] = 9
df.loc[df.index[df['INF3isscare']==0], 'INF3isscare'] = 9
df.loc[df.index[df['INF3medical']==0], 'INF3medical'] = 9
df.loc[df.index[df['INF3meds']==0], 'INF3meds'] = 9

df.loc[df.index[df['INF3breast']==0], 'INF3breast'] = 9

df.loc[df.index[df['INF3breastatdeath']==0], 'INF3breastatdeath'] = 9

################################################################
################################################################ smoking cigarettes
################################################################
## combine zeros and nines: blanks and unknowns
df.loc[df.index[df['INF3smbefpreg']==0],'INF3smbefpreg']=9
df.loc[df.index[df['INF3smanytime']==0],'INF3smanytime']=9



'''
Can you make a variable that is smoked before pregnancy OR first trimester 
OR second trimester OR third trimester (i.e. categorise as Smoker)? 
Categorise non-smoker if No at any time if not defined as a smoker, 
otherwise categorise as Unknown. I think that will be the best we can do.

'''
df.loc[df.index[df['INF3firsttricig']>1], 'INF3smanytime'] = 1
df.loc[df.index[df['INF3sectricig']>1], 'INF3smanytime'] = 1
df.loc[df.index[df['INF3thirdtricig']>1], 'INF3smanytime'] = 1

df['Smoker']=1
for i in df.index:
	subject=df.iloc[i]
	if subject['INF3smbefpreg']==9 and subject['INF3smanytime']==9: df.loc[i, 'Smoker']=9
	if subject['INF3smbefpreg']==9 and subject['INF3smanytime']==2: df.loc[i, 'Smoker']=2

	if subject['INF3smbefpreg']==2 and subject['INF3smanytime']==9: df.loc[i, 'Smoker']=2
	if subject['INF3smbefpreg']==2 and subject['INF3smanytime']==2: df.loc[i, 'Smoker']=2
	
	if subject['INF3smbefpreg']==9 and subject['INF3smanytime']==1: df.loc[i, 'Smoker']=1
	if subject['INF3smbefpreg']==2 and subject['INF3smanytime']==1: df.loc[i, 'Smoker']=1

	if subject['INF3smbefpreg']==1 and subject['INF3smanytime']==1: df.loc[i, 'Smoker']=1
	if subject['INF3smbefpreg']==1 and subject['INF3smanytime']==2: df.loc[i, 'Smoker']=1
	if subject['INF3smbefpreg']==1 and subject['INF3smanytime']==9: df.loc[i, 'Smoker']=1


c1 = df.iloc[df.index[df['INF3smbefpreg']==1]]
c2 = df.iloc[df.index[df['INF3smbefpreg']==2]]
c9 = df.iloc[df.index[df['INF3smbefpreg']==9]]

print('c1')
print(len(c1.index[c1['INF3smanytime']==1]))
print(len(c1.index[c1['INF3smanytime']==2]))
print(len(c1.index[c1['INF3smanytime']==9]))

len(c1.index[c1['INF3smanytime']==1])+len(c1.index[c1['INF3smanytime']==2])+len(c1.index[c1['INF3smanytime']==9])
len(c2.index[c2['INF3smanytime']==1])+len(c2.index[c2['INF3smanytime']==2])+len(c2.index[c2['INF3smanytime']==9])
len(c9.index[c9['INF3smanytime']==1])+len(c9.index[c9['INF3smanytime']==2])+len(c9.index[c9['INF3smanytime']==9])

print('c2')
print(len(c2.index[c2['INF3smanytime']==1]))
print(len(c2.index[c2['INF3smanytime']==2]))
print(len(c2.index[c2['INF3smanytime']==9]))

print('c3')
print(len(c9.index[c9['INF3smanytime']==1]))
print(len(c9.index[c9['INF3smanytime']==2]))
print(len(c9.index[c9['INF3smanytime']==9]))

df['INF3tobacco'] = df['Smoker']
# df.loc[df.index[df['INF3smanytime']==9], 'INF3tobacco'] = 9
# df.loc[df.index[df['INF3smbefpreg']==9], 'INF3tobacco'] = 9
# df.loc[df.index[df['INF3smbefpreg']==1], 'INF3tobacco'] = 1
# df.loc[df.index[df['INF3smanytime']==1], 'INF3tobacco'] = 1

# df.loc[df.index[df['INF3firsttricig']>0], 'INF3tobacco'] = 1
# df.loc[df.index[df['INF3firsttricigunk']==1], 'INF3tobacco'] = 1
# # df.loc[df.index[df['INF3firsttricig'].isna()], 'INF3tobacco'] = 9

# df.loc[df.index[df['INF3sectricig']>0], 'INF3tobacco'] = 1
# df.loc[df.index[df['INF3sectricigunk']==1], 'INF3tobacco'] = 1
# # df.loc[df.index[df['INF3sectricig'].isna()], 'INF3tobacco'] = 9

# df.loc[df.index[df['INF3thirdtricig']>0], 'INF3tobacco'] = 1
# df.loc[df.index[df['INF3thirdtricigunk']==1], 'INF3tobacco'] = 1
# # df.loc[df.index[df['INF3thirdtricig'].isna()], 'INF3tobacco'] = 9

# df.loc[df.index[df['INF3firsttricig']==0], 'INF3firsttricig'] = np.nan
# df.loc[df.index[df['INF3sectricig']==0], 'INF3sectricig'] = np.nan
# df.loc[df.index[df['INF3thirdtricig']==0], 'INF3thirdtricig'] = np.nan
################
df.loc[df.index[df['GIV1type']==0], 'GIV1type'] = 99
df.loc[df.index[df['GIV1age']==0], 'GIV1ageunk'] = 1
df.loc[df.index[df['GIV1age']==1], 'GIV1age'] = np.nan
df.loc[df.index[df['GIV1age']==0], 'GIV1age'] = np.nan
df.loc[df.index[df['GIV1sex']==0], 'GIV1sex'] =9
df.loc[df.index[df['GIV1employ']==0],'GIV1employ']=9
df.loc[df.index[df['GIV1income']==0], 'GIV1income'] =9
df.loc[df.index[df['GIV1educt']==0],'GIV1educt']=9
df.loc[df.index[df['GIV1SocialServices']==0],'GIV1SocialServices']=9
df.loc[df.index[df['GIV1drugabuse']==0],'GIV1drugabuse']=9

# df.loc[df.index[df['GIV1age'].isna()],'GIV1age']= 1
df.loc[df.index[df['GIV1sex'].isna()],'GIV1sex']= 9
df.loc[df.index[df['GIV1employ'].isna()],'GIV1employ']= 9
df.loc[df.index[df['GIV1income'].isna()],'GIV1income']= 9
df.loc[df.index[df['GIV1educt'].isna()],'GIV1educt']= 9
df.loc[df.index[df['GIV1SocialServices'].isna()],'GIV1SocialServices']= 9
df.loc[df.index[df['GIV1drugabuse'].isna()],'GIV1drugabuse']= 9

df.loc[df.index[df['GIV2type']==0], 'GIV2type'] = 99
df.loc[df.index[df['GIV2age']==0], 'GIV2ageunk'] = 1
df.loc[df.index[df['GIV2age']==1], 'GIV2age'] = np.nan
df.loc[df.index[df['GIV2age']==0], 'GIV2age'] = np.nan
df.loc[df.index[df['GIV2sex']==0], 'GIV2sex'] =9
df.loc[df.index[df['GIV2employ']==0],'GIV2employ']=9
df.loc[df.index[df['GIV2income']==0], 'GIV2income'] =9
df.loc[df.index[df['GIV2educt']==0],'GIV2educt']=9
df.loc[df.index[df['GIV2SocialServices']==0],'GIV2SocialServices']=9
df.loc[df.index[df['GIV2drugabuse']==0],'GIV2drugabuse']=9

# df.loc[df.index[df['GIV2age'].isna()], 'GIV2age'] = 1
# df.loc[df.index[df['GIV2age']==1], 'GIV2age'] = np.nan
df.loc[df.index[df['GIV2sex'].isna()], 'GIV2sex'] = 9
df.loc[df.index[df['GIV2employ'].isna()], 'GIV2employ'] = 9
df.loc[df.index[df['GIV2income'].isna()], 'GIV2income'] = 9
df.loc[df.index[df['GIV2educt'].isna()], 'GIV2educt'] = 9
df.loc[df.index[df['GIV2SocialServices'].isna()], 'GIV2SocialServices'] = 9
df.loc[df.index[df['GIV2drugabuse'].isna()], 'GIV2drugabuse'] = 9

df.loc[df.index[df['GIV2age']>=99], 'GIV2age'] = np.nan

### eliminate nonsense age, no caregiver could be younger than 10 years of age:
df.loc[df.index[df['GIV1age']<10], 'GIV1age']=np.nan
df.loc[df.index[df['GIV2age']<10], 'GIV1age']=np.nan

######################## combine CAREGIVERS 

caregiver11 = df.index[df['GIV1type']==21]
caregiver12 = df.index[df['GIV1type']==22]
caregiver01 = df.index[df['GIV1type']==99].tolist()

caregiver21 = df.index[df['GIV2type']==21]
caregiver22 = df.index[df['GIV2type']==22]
caregiver02 = df.index[df['GIV2type']==99].tolist()

caregiver01 = set(caregiver01)
caregiver02 = set(caregiver02)
who_raised_this_baby = list(caregiver01.intersection(caregiver02))

df['caregiver'] = 2
df.loc[who_raised_this_baby, 'caregiver'] = 9
df.loc[caregiver11, 'caregiver'] = 1
df.loc[caregiver12, 'caregiver'] = 1
df.loc[caregiver21, 'caregiver'] = 1
df.loc[caregiver21, 'caregiver'] = 1

### encoded: drug abuse
# df.loc[df.index[df['GIV1drugabuse']==1], 'GIVs_drugs'] = 1
# df.loc[df.index[df['GIV2drugabuse']==1], 'GIVs_drugs'] = 1

# clean = df.iloc[df.index[df['GIV1drugabuse']==2]]
# df.loc[clean.index[clean['GIV2drugabuse']==2], 'GIVs_drugs'] =2
# df.loc[clean.index[clean['GIV2drugabuse']==9], 'GIVs_drugs'] =2
# cleann = df.iloc[df.index[df['GIV2drugabuse']==2]]
# df.loc[cleann.index[cleann['GIV1drugabuse']==9], 'GIVs_drugs']=2

# blank = df.iloc[df.index[df['GIV1drugabuse']==9]]
# df.loc[blank.index[blank['GIV2drugabuse']==9], 'GIVs_drugs'] = 9

df['GIVs_drugs']=9
for idx in df.index:
	row=df.iloc[idx]
	##1
	if row['GIV1drugabuse']==1:
		df.loc[idx, 'GIVs_drugs']=1
	##2
	if row['GIV2drugabuse']==1:
		df.loc[idx, 'GIVs_drugs']=1
	##3
	if row['GIV1drugabuse']==2:
		if row['GIV2drugabuse']==2:
			df.loc[idx, 'GIVs_drugs']=2
		if row['GIV2drugabuse']==9:
			df.loc[idx, 'GIVs_drugs']=2
	if row['GIV2drugabuse']==2:
		if row['GIV1drugabuse']==9:
			df.loc[idx, 'GIVs_drugs']=2
	


### education
df.loc[df.index[df['GIV1educt']==9], 'GIV1educt']=0
df.loc[df.index[df['GIV2educt']==9], 'GIV2educt']=0
df['GIVs_edu']=9

for idx in df.index:
	row=df.iloc[idx]
	parent_edu_score = np.array([row['GIV1educt'],row['GIV2educt']])
	df.loc[idx, 'GIVs_edu']=parent_edu_score.max()


# df.loc[df.index[df['GIV1educt']==1], 'GIVs_edu'] = 1
# df.loc[df.index[df['GIV2educt']==1], 'GIVs_edu'] = 1
# df.loc[df.index[df['GIV1educt']==2], 'GIVs_edu'] = 2
# df.loc[df.index[df['GIV2educt']==2], 'GIVs_edu'] = 2
# df.loc[df.index[df['GIV1educt']==3], 'GIVs_edu'] = 3
# df.loc[df.index[df['GIV2educt']==3], 'GIVs_edu'] = 3
# df.loc[df.index[df['GIV1educt']==4], 'GIVs_edu'] = 4
# df.loc[df.index[df['GIV2educt']==4], 'GIVs_edu'] = 4

######## income
df['GIVs_income'] = 9
df.loc[df.index[df['GIV1income']==3], 'GIVs_income'] = 3
df.loc[df.index[df['GIV2income']==3], 'GIVs_income'] = 3
df.loc[df.index[df['GIV1income']==2], 'GIVs_income'] = 2
df.loc[df.index[df['GIV2income']==2], 'GIVs_income'] = 2
df.loc[df.index[df['GIV1income']==1], 'GIVs_income'] = 1
df.loc[df.index[df['GIV2income']==1], 'GIVs_income'] = 1

########## Social Services
df['GIVs_ss'] = 9
df.loc[df.index[df['GIV1SocialServices']==2], 'GIVs_ss'] = 2
df.loc[df.index[df['GIV2SocialServices']==2], 'GIVs_ss'] = 2
df.loc[df.index[df['GIV1SocialServices']==1], 'GIVs_ss'] = 1
df.loc[df.index[df['GIV2SocialServices']==1], 'GIVs_ss'] = 1

#########################################################################
df.loc[df.index[df['INCarea']==0],'INCarea']=9
df.loc[df.index[df['INCothplacesp']==''],'INCothplacesp'] = np.nan
# df.loc[df.index[df['INCplace']==0], 'INCplace']=9

df.loc[df.index[df['CAUextinjury']==0],'CAUextinjury']=99

p1=df.iloc[df.index[df['INCchhome']==1]]
p2=df.iloc[df.index[df['INClicdayhm']==1]]
p3=df.iloc[df.index[df['INCothplace']==1]]

p1 = set(p1.index.tolist())
p2 = set(p2.index.tolist())
p3 = set(p3.index.tolist())
int1 = p1.intersection(p2)
int2 = p2.intersection(p3)
int3 = p3.intersection(p1)

# print('Not mutually exclusive')

df['INCplace'] = 9
df.loc[df.index[df['INCothplace']==1], 	'INCplace'] = 3
df.loc[df.index[df['INClicdayhm']==1], 	'INCplace'] = 2
df.loc[df.index[df['INCchhome']==1], 	'INCplace'] = 1
k1= df.iloc[df.index[df['INCchhome']==0]]
#
# p055=validate()
# df.loc[df.index[df['CAUextinjury']==0],'CAUextinjury']=np.nan
# df.loc[df.index[df['CAUmedcond']==0],'CAUmedcond']=np.nan

df.loc[df.index[df['SUFevent']==0],'SUFevent']=99
df.loc[df.index[df['SUFcause']==0],'SUFcause']=99
#
####################################################################
df.loc[df.index[df['CIRtosleep']>=1], 'CIRdeathsleepenv'] = 1
# df.loc[df.index[df['CIRtosleep']>=1], 'CIRdeathsleepenv'] = 1
df.loc[df.index[df['CIRdeathsleepenv']!=1], 'CIRdeathsleepenv'] = 0
df.loc[df.index[df['CIRsleepplace']==0],'CIRsleepplace']=99
df.loc[df.index[df['CIRtosleep']==0],'CIRtosleep']=9
df.loc[df.index[df['CIRfoundpos']==0],'CIRfoundpos']=9
df.loc[df.index[df['CIRuslplace']==0],'CIRuslplace']=99
df.loc[df.index[df['CIRuslslppos']==0],'CIRuslslppos']=9
df.loc[df.index[df['CIRnewenvir']==0],'CIRnewenvir']=9
df.loc[df.index[df['CIRpacifier']==0],'CIRpacifier']=9
df.loc[df.index[df['CIRswaddled']==0],'CIRswaddled']=9
df.loc[df.index[df['CIRoverheat']==0],'CIRoverheat']=9

df.loc[df.index[df['CIRroomhot']!=1],'CIRroomhot']=0
df.loc[df.index[df['CIRtoobed']!=1],'CIRtoobed']=0
df.loc[df.index[df['CIRtoocloth']!=1],'CIRtoocloth']=0
df.loc[df.index[df['CIRsecsm']==0],'CIRsecsm']=9
df.loc[df.index[df['CIRsmokfreq']==0],'CIRsmokfreq']=9
df.loc[df.index[df['CIRfacepos']==0],'CIRfacepos']=9
df.loc[df.index[df['CIRpresadult']==0],'CIRpresadult']=9
# df.loc[df.index[df['CIRpresadultsp']==0],'CIRpresadultsp']=9
df.loc[df.index[df['CIRpreschild']==0],'CIRpreschild']=9
df.loc[df.index[df['CIRpresaniml']==0],'CIRpresaniml']=9

df.loc[df.index[df['CIRpresmattrss']==0],'CIRpresmattrss']=9
df.loc[df.index[df['CIRprescomfrtr']==0],'CIRprescomfrtr']=9
df.loc[df.index[df['CIRpresfitsheet']==0],'CIRpresfitsheet']=9
df.loc[df.index[df['CIRpresblnkt']==0],'CIRpresblnkt']=9
df.loc[df.index[df['CIRprespillow']==0],'CIRprespillow']=9
df.loc[df.index[df['CIRprescushion']==0],'CIRprescushion']=9
df.loc[df.index[df['CIRpresboppy']==0],'CIRpresboppy']=9
df.loc[df.index[df['CIRprespositioner']==0],'CIRprespositioner']=9
df.loc[df.index[df['CIRpresbmppad']==0],'CIRpresbmppad']=9
df.loc[df.index[df['CIRpresclothing']==0],'CIRpresclothing']=9

df.loc[df.index[df['CIRprescribrail']==0],'CIRprescribrail']=9
df.loc[df.index[df['CIRpreswall']==0],'CIRpreswall']=9
df.loc[df.index[df['CIRprestoy']==0],'CIRprestoy']=9
df.loc[df.index[df['CIRpresother1']!=1],'CIRpresother1']=0
## df.loc[df.index[df['CIRpresother1sp']==0],'CIRpresother1sp']=9
df.loc[df.index[df['CIRpresother2']!=1],'CIRpresother2']=0
## df.loc[df.index[df['CIRpresother2sp']==0],'CIRpresother2sp']=9
df.loc[df.index[df['CIRadltfeeding']==0],'CIRadltfeeding']=9
df.loc[df.index[df['CIRroomshare']==0],'CIRroomshare']=9
df.loc[df.index[df['CIRsamesurf']==0],'CIRsamesurf']=9

df.loc[df.index[df['CIRwadult']!=1],'CIRwadult']=0
df.loc[df.index[df['CIRwchild']!=1],'CIRwchild']=0
df.loc[df.index[df['CIRwpet']!=1],'CIRwpet']=0

df['GIV1race'] = 0
df.loc[df.index[df['GIV1raceai']==1], 'GIV1race'] = 1
df.loc[df.index[df['GIV1raceak']==1], 'GIV1race'] = 1
df.loc[df.index[df['GIV1raceas']==1], 'GIV1race'] = 2
df.loc[df.index[df['GIV1racebl']==1], 'GIV1race'] = 3
df.loc[df.index[df['GIV1racehi']==1], 'GIV1race'] = 4
df.loc[df.index[df['GIV1racepi']==1], 'GIV1race'] = 4
df.loc[df.index[df['GIV1racewh']==1], 'GIV1race'] = 5

df['GIV2race'] = 0
df.loc[df.index[df['GIV2raceai']==1], 'GIV2race'] = 1
df.loc[df.index[df['GIV2raceak']==1], 'GIV2race'] = 1
df.loc[df.index[df['GIV2raceas']==1], 'GIV2race'] = 2
df.loc[df.index[df['GIV2racebl']==1], 'GIV2race'] = 3
df.loc[df.index[df['GIV2racehi']==1], 'GIV2race'] = 4
df.loc[df.index[df['GIV2racepi']==1], 'GIV2race'] = 4
df.loc[df.index[df['GIV2racewh']==1], 'GIV2race'] = 5
#################################################################

#################################
# df.to_csv('data/NFCRP_SeaChi.csv')
#################################################################
# tab = df.copy()

# tab['INFdodweek'] = df['INFdodweek'].copy()
# tab['INFdodyear'] = df['INFdodyear'].copy()
# tab['INFethnic'] = df['INFethnic'].copy()
# tab['INFsex'] = df['INFsex'].copy()
# tab['INFageMonths'] = df['INFageMonths'].copy()
# tab['INF3gestage'] = df['INF3gestage'].copy()

# tab['INF3multbirth'] = df['INF3multbirth'].copy()
# tab['INF3prenatalcare'] = df['INF3prenatalcare'].copy()
# # tab['INF3prenatal'] = df['INF3prenatal'].copy()
# tab['INF3vismonth'] = df['INF3vismonth'].copy()
# tab['INF3medical'] = df['INF3medical'].copy()

# tab['INF3tobacco'] = df['INF3tobacco'].copy()
# tab['INF3breast'] = df['INF3breast'].copy()

# tab['GIVs_drugs'] = df['GIVs_drugs'].copy()
# tab['GIVs_edu'] = df['GIVs_edu'].copy()
# tab['GIVs_income'] = df['GIVs_income'].copy()

# tab['GIV1type'] = df['GIV1type'].copy()
# tab['GIV1age'] = df['GIV1age'].copy()
# tab['GIV1sex'] = df['GIV1sex'].copy()
# tab['GIV1educt'] = df['GIV1educt'].copy()
# tab['GIV1race'] = df['GIV1race'].copy()

# tab['GIV2type'] = df['GIV2type'].copy()
# tab['GIV2age'] = df['GIV2age'].copy()
# tab['GIV2sex'] = df['GIV2sex'].copy()
# tab['GIV2educt'] = df['GIV2educt'].copy()
# tab['GIV2race'] = df['GIV2race'].copy()

# tab['INCarea'] = df['INCarea'].copy()
# tab['INCplace'] = df['INCplace'].copy()

# tab['CIRdeathsleepenv'] = df['CIRdeathsleepenv'].copy()
# tab['CIRsleepplace'] = df['CIRsleepplace'].copy()
# tab['CIRtosleep'] = df['CIRtosleep'].copy()
# tab['CIRfoundpos'] = df['CIRfoundpos'].copy()

# tab['CIRnewenvir'] = df['CIRnewenvir'].copy()
# tab['CIRpresadult'] = df['CIRpresadult'].copy()
# tab['CIRpreschild'] = df['CIRpreschild'].copy()
# tab['CIRpresaniml'] = df['CIRpresaniml'].copy()

# tab['CIRadltfeeding'] = df['CIRadltfeeding'].copy()
# tab['CIRroomshare'] = df['CIRroomshare'].copy()
# tab['CIRwadult'] = df['CIRwadult'].copy()
# tab['CIRwchild'] = df['CIRwchild'].copy()
# tab['CIRwpet'] = df['CIRwpet'].copy()
# tab['CAURcause'] = df['CAURcause'].copy()
# tab['INFRrace_AI'] = df['INFRrace_AI'].copy()

# tab['CIRpresblnkt'] = df['CIRpresblnkt'].copy()
# tab['CIRprescushion'] = df['CIRprescushion'].copy()
# tab['CIRpresboppy'] = df['CIRpresboppy'].copy()
# tab['CIRpresclothing'] = df['CIRpresclothing'].copy()
# tab['CIRprestoy'] = df['CIRprestoy'].copy()
###################################################################

# df['CIR_cosleep']= '?'
# df.loc[df.index[df['CIRsleepplace']==1], 'CIR_cosleep'] = 'solitary'
# df.loc[df.index[df['CIRsleepplace']==2], 'CIR_cosleep'] = 'solitary'
# df.loc[df.index[df['CIRsleepplace']==5], 'CIR_cosleep'] = 'solitary'
# df.loc[df.index[df['CIRsleepplace']==9], 'CIR_cosleep'] = 'solitary'
# df.loc[df.index[df['CIRsleepplace']==10], 'CIR_cosleep'] = 'solitary'
# df.loc[df.index[df['CIRsleepplace']==13], 'CIR_cosleep'] = 'solitary'
# df.loc[df.index[df['CIRsleepplace']==14], 'CIR_cosleep'] = 'solitary'
# df.loc[df.index[df['CIRsleepplace']==15], 'CIR_cosleep'] = 'solitary'
# df.loc[df.index[df['CIRsleepplace']==16], 'CIR_cosleep'] = 'solitary'
# df.loc[df.index[df['CIRsleepplace']==17], 'CIR_cosleep'] = 'solitary'

# anomalies = df.loc[df.index[df['CIR_cosleep']=='solitary']]
# cases = df.loc[anomalies.index[anomalies['CIRsamesurf']==1]]


# j=0
# for i in eds_handbook.index:
# 	row = eds_handbook.iloc[i]
# 	if row['Bed_sharing'] == 'bed sharing' or row['Bed_sharing'] == 'Bed sharing':
# 		# print(i)
# 		df.loc[df.index[df['CAUicdcode']==row['CAUicdcode']], 'CIR_cosleep'] = 'yes'

# 	if row['Eds_Code']=='W75':
# 		df.loc[df.index[df['CAUicdcode']==row['CAUicdcode']], 'CIR_cosleep'] = 'yes'


# tab['CIRsamesurf'] = df['CIRsamesurf'].copy()
# print('Before:\n',tab['CIRsamesurf'].value_counts(),'\n')

# df.loc[df.index[df['CIR_cosleep']=='yes'], 'CIRsamesurf']= 1
# df.loc[df.index[df['CIR_cosleep']=='solitary'], 'CIRsamesurf']= 2

# tab['CIRsamesurf'] = df['CIRsamesurf'].copy()
# # print('After:\n',tab['CIRsamesurf'].value_counts())
# # 



###########################################################################################
#########################################
# Bedsharing=1 cannot be overridden. Bedsharing=2 can move to 1. Bedsharing=3 can move to 1, 2 or remain at 9.
#1
df['Bed_sharing']=df['CIRsamesurf'].copy()
t1=select_sample(df)
print('\n1:\n',df['Bed_sharing'].value_counts(),t1['Bed_sharing'].value_counts(),'\n')
#2
# m=[]
# n=[]
# for i in eds_handbook.index:
# 	row = eds_handbook.iloc[i]
# 	if row['Bed_sharing'] == 'bed sharing' or row['Bed_sharing'] == 'Bed sharing':
# 		# print(i)
# 		df.loc[df.index[df['CAUicdcode']==row['CAUicdcode']], 'Bed_sharing'] = 1
# 		m.append(df.index[df['CAUicdcode']==row['CAUicdcode']])

# t2=select_sample(df)
# print('2:',df['Bed_sharing'].value_counts(),t2['Bed_sharing'].value_counts(),'\n')

# for i in eds_handbook.index:
# 	row = eds_handbook.iloc[i]
# #3
# 	if row['Eds_Code']=='W75':
# 		df.loc[df.index[df['CAUicdcode']==row['CAUicdcode']], 'Bed_sharing'] = 1
# 		n.append(df.index[df['CAUicdcode']==row['CAUicdcode']])
# t3=select_sample(df)
# print('3:',df['Bed_sharing'].value_counts(),t3['Bed_sharing'].value_counts(),'\n')


# #3.1 if Bed_sharing =1 got reassigned, fix:
# j=0
# for i in df.index:
# 	row=df.iloc[i]
# 	if row['CIRsamesurf']==1:
# 		if row['Bed_sharing']!=1:
# 			df.loc[i, 'Bed_sharing']=1
# 			j+=1
####### current output: j=0; correction not needed. But implemented to certify that Bed_sharing=1 never gets reassigned.
#4
df['CIR_cosleep']='?'
df.loc[df.index[df['CIRsleepplace']==1], 'CIR_cosleep'] = 'solitary'
df.loc[df.index[df['CIRsleepplace']==2], 'CIR_cosleep'] = 'solitary'
df.loc[df.index[df['CIRsleepplace']==5], 'CIR_cosleep'] = 'solitary'
df.loc[df.index[df['CIRsleepplace']==9], 'CIR_cosleep'] = 'solitary'
df.loc[df.index[df['CIRsleepplace']==10], 'CIR_cosleep'] = 'solitary'
df.loc[df.index[df['CIRsleepplace']==13], 'CIR_cosleep'] = 'solitary'
df.loc[df.index[df['CIRsleepplace']==14], 'CIR_cosleep'] = 'solitary'
df.loc[df.index[df['CIRsleepplace']==15], 'CIR_cosleep'] = 'solitary'
df.loc[df.index[df['CIRsleepplace']==16], 'CIR_cosleep'] = 'solitary'
df.loc[df.index[df['CIRsleepplace']==17], 'CIR_cosleep'] = 'solitary'

k=0
#items from 4
z=[]
for i in df.index:
	row=df.iloc[i]
	if row['Bed_sharing']==9:
		if row['CIR_cosleep']=='solitary':
			df.loc[i,'Bed_sharing']=2
			k+=1
			if row['INFdodyear']>2014:
				z.append(i)
t4=select_sample(df)
print('4:',df['Bed_sharing'].value_counts(),t4['Bed_sharing'].value_counts(),'\n')


# print(f'Items fixed in 2:{np.sum(m)}\nItems fixed in 3:{np.sum(n)}\nItems corrected back to 1(3.1):{j}\nItems fixed in 4:{k}')
# print(str(np.sum(m)+np.sum(n)+j+k))

# tab['Bed_sharing']=df['Bed_sharing'].copy()
# tab['CIRsamesurf']=df['CIRsamesurf'].copy()

# ## items from 2
# y=[]
# for a in m:
# 	for b in a:
# 		if df.iloc[b]['INFdodyear']>2014:
# 			if df.iloc[b]['CIRsamesurf']!=1:
# 				y.append(b)	

# ## items from 3
# x=[]
# for a in n:
# 	for b in a:
# 		if df.iloc[b]['INFdodyear']>2014:
# 			if df.iloc[b]['CIRsamesurf']!=1:
# 				x.append(b)


# x_ = df.iloc[x]
# y_ = df.iloc[y]
z_ = df.iloc[z]
# print('What was corrected:\nFrom 2:')
# print(y_['CIRsamesurf'].value_counts())
# print('From 3:')
# print(x_['CIRsamesurf'].value_counts())
print('From 4:')
print(z_['CIRsamesurf'].value_counts())

###########################################################################################
################ recode birthweight:
df['INF3biwt_']=''
df.loc[df.index[df['INF3biwt']<2500], 'INF3biwt_'] = '<2500'

def recode_birthweight(min_,max_):
	u = df.index[df['INF3biwt']<=max_]
	uu=set(df.index[u])
	v = df.index[df['INF3biwt']>=min_]
	vv=set(df.index[v])

	# prime = uu.symmetric_difference(vv)
	prime = uu.intersection(vv)

	df.loc[list(prime), 'INF3biwt_'] = f'{min_}-{max_}'

# recode_birthweight(1000,1499)
# recode_birthweight(1500,1999)
# recode_birthweight(2000,2499)

recode_birthweight(2500,2999)
recode_birthweight(3000,3499)
recode_birthweight(3500,4000)

df.loc[df.index[df['INF3biwt']>4000], 'INF3biwt_'] = '+4000'

# tab['INF3biwt'] = df['INF3biwt'].copy()
# tab['INF3biwt_'] = df['INF3biwt_'].copy()



########## variables of interest: PEAKED EDS INTEREST BUT PERCETAGE OF MISSING DATA 
######### for forecasting:

# tab['CIRuslplace'] = df['CIRuslplace'].copy()
# tab['CIRuslslppos'] = df['CIRuslslppos'].copy()

# tab['CIRpacifier'] = df['CIRpacifier'].copy()
# tab['CIRswaddled'] = df['CIRswaddled'].copy()
# tab['CIRsecsm'] = df['CIRsecsm'].copy()
# tab['CIRfacepos'] = df['CIRfacepos'].copy()

# tab['CIRprescomfrtr'] = df['CIRprescomfrtr'].copy()
# tab['CIRprespillow'] = df['CIRprespillow'].copy()

# tab['CAUcausedth'] = df['CAUcausedth'].copy()

# tab['caregiver'] = df['caregiver'].copy()

# tab['Smoker'] = df['Smoker'].copy()



##################################### save as a csv export in my working directory.
# tab['INF3biwt'] = df['INF3biwt'].copy()
# tab['INF3biwt_'] = df['INF3biwt_'].copy()

# tab['INFageDays'] = df['INFageDays'].copy()

tab = df.copy()

####

tab=select_sample(tab)
print('\nFinal tables:',tab['CIRsamesurf'].value_counts(),'\n',tab['Bed_sharing'].value_counts())
##
tab.to_csv('data/NCFRP_SCH_sample.csv')
df.to_csv('data/NCFRP_SCH_pop.csv')



























